from typing import Dict, List

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - openpyxl pode não estar disponível
    Workbook = None
    load_workbook = None

PROFISSIONAIS = [
    {
        "nome": "Fulano",
        "titulo": "Desenvolvedor Python",
        "pais": "Brasil",
        "estado": "SP",
        "cidade": "Sao Paulo",
        "setor": "Tecnologia",
        "senioridade": "Sênior",
        "skills": ["Python", "Django", "AWS"],
        "empresa": "Empresa X",
        "formacao": "Bacharelado",
        "certificacoes": ["AWS"],
    },
    {
        "nome": "Beltrano",
        "titulo": "Product Manager",
        "pais": "Brasil",
        "estado": "RJ",
        "cidade": "Rio de Janeiro",
        "setor": "Tecnologia",
        "senioridade": "Pleno",
        "skills": ["Agile", "Scrum"],
        "empresa": "Empresa Y",
        "formacao": "MBA",
        "certificacoes": ["PMP"],
    },
    {
        "nome": "Sicrano",
        "titulo": "Advogado",
        "pais": "Brasil",
        "estado": "MG",
        "cidade": "Belo Horizonte",
        "setor": "Jurídico",
        "senioridade": "Sênior",
        "skills": ["Direito", "Contratos"],
        "empresa": "Escritorio Z",
        "formacao": "Mestrado",
        "certificacoes": [],
    },
    {
        "nome": "Maria",
        "titulo": "Engenheira de Software",
        "pais": "Brasil",
        "estado": "MG",
        "cidade": "Belo Horizonte",
        "setor": "Tecnologia",
        "senioridade": "Júnior",
        "skills": ["Python", "UX"],
        "empresa": "Empresa K",
        "formacao": "Bacharelado",
        "certificacoes": [],
    },
]


def _gerar_profissionais_teste():
    estados = ["SP", "RJ", "MG", "RS", "SC"]
    cidades = {
        "SP": ["Sao Paulo", "Campinas", "Santos"],
        "RJ": ["Rio de Janeiro", "Niteroi"],
        "MG": ["Belo Horizonte", "Uberlandia"],
        "RS": ["Porto Alegre", "Caxias do Sul"],
        "SC": ["Florianopolis", "Joinville"],
    }
    setores = ["Tecnologia", "Finanças", "Saúde", "Educação", "Jurídico"]
    senioridades = ["Estágio", "Júnior", "Pleno", "Sênior", "Lead"]
    titulos = [
        "Analista de Teste",
        "Desenvolvedor Backend",
        "Desenvolvedor Frontend",
        "Analista de Dados",
        "Designer",
    ]
    formacoes = [
        "Bacharelado",
        "Mestrado",
        "MBA",
        "Tecnólogo",
        "Especialização",
    ]
    certificacoes_opcoes = [
        [],
        ["AWS"],
        ["PMP"],
        ["Scrum Master"],
        ["ITIL"],
        ["CCNA"],
        ["Docker", "Kubernetes"],
        ["Azure"],
    ]

    for i in range(100):
        estado = estados[i % len(estados)]
        cidade = cidades[estado][i % len(cidades[estado])]
        PROFISSIONAIS.append(
            {
                "nome": f"Teste {i}",
                "titulo": titulos[i % len(titulos)],
                "pais": "Brasil",
                "estado": estado,
                "cidade": cidade,
                "setor": setores[i % len(setores)],
                "senioridade": senioridades[i % len(senioridades)],
                "skills": ["Teste", f"Skill{i}"],
                "empresa": f"Empresa Teste {i}",
                "formacao": formacoes[i % len(formacoes)],
                "certificacoes": certificacoes_opcoes[i % len(certificacoes_opcoes)],
            }
        )


_gerar_profissionais_teste()


# ---------------------------------------------------------------------------
# Utilidades para arquivo Excel
# ---------------------------------------------------------------------------


REQUIRED_COLUMNS = [
    "nome",
    "titulo",
    "pais",
    "estado",
    "cidade",
    "setor",
    "senioridade",
    "skills",
    "empresa",
    "formacao",
    "certificacoes",
]


def salvar_excel(path: str) -> None:
    """Salva um arquivo Excel com os 100 primeiros registros fictícios."""

    if Workbook is None:  # pragma: no cover - dependencia opcional
        raise ImportError("openpyxl não está disponível")

    wb = Workbook()  # type: ignore[call-arg]
    ws = wb.active  # type: ignore[assignment]
    ws.append(REQUIRED_COLUMNS)  # type: ignore[operator]

    for p in PROFISSIONAIS[:100]:
        row = []
        for col in REQUIRED_COLUMNS:
            val = p.get(col, "")
            if isinstance(val, list):
                val = ", ".join(val)
            row.append(val)
        ws.append(row)  # type: ignore[operator]

    wb.save(path)  # type: ignore[operator]


def carregar_excel(path: str) -> List[Dict]:
    """Carrega profissionais a partir de um arquivo Excel."""

    if load_workbook is None:  # pragma: no cover - dependencia opcional
        raise ImportError("openpyxl não está disponível")

    wb = load_workbook(path)  # type: ignore[call-arg]
    ws = wb.active  # type: ignore[assignment]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]  # type: ignore[operator]
    if not all(col in headers for col in REQUIRED_COLUMNS):
        raise ValueError("colunas ausentes")

    indices = {h: headers.index(h) for h in REQUIRED_COLUMNS}
    dados: List[Dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore[operator]
        registro = {}
        for col in REQUIRED_COLUMNS:
            idx = indices[col]
            val = row[idx] if idx < len(row) else None
            if col in {"skills", "certificacoes"}:
                if isinstance(val, str):
                    val = [v.strip() for v in val.split(",") if v.strip()]
                elif val is None:
                    val = []
                else:
                    val = list(val) if isinstance(val, (list, tuple)) else [val]
            registro[col] = val
        dados.append(registro)

    return dados
